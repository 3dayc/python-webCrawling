from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import bs4
import openpyxl as xl

base_url = "https://www.datacamp.com"
url =f"{base_url}/tracks/machine-learning-scientist-with-python"

driver = webdriver.Chrome()
driver.implicitly_wait(3)
driver.get(url)

wait = WebDriverWait(driver, 10)
wait.until(EC.element_to_be_clickable((By.XPATH, """//*[@id="gatsby-focus-wrapper"]/div/div[1]/div[1]/div/div/div[4]/button""")))

btn = driver.find_element_by_xpath("""//*[@id="gatsby-focus-wrapper"]/div/div[1]/div[1]/div/div/div[4]/button""")
btn.click()

bs = bs4.BeautifulSoup(driver.page_source, features="html.parser")

courses = bs.select("#gatsby-focus-wrapper > div > div.container.css-93pq91 > div.col-md-8 > div > div > div > div.css-10s95pl > a")
courseList = []
for c in courses:
    link = c.attrs["href"]
    title = c.select_one("h4").getText().strip()
    desc = c.select_one("p").getText().strip()
    courseList.append({"link": link, "title": title, "desc": desc})

for c in courseList:
    driver.get(f"{base_url}{c['link']}")
    
    bs_detail = bs4.BeautifulSoup(driver.page_source, features="html.parser")

    chapters = bs_detail.select_one("ol.chapters")

    chapters_elem = chapters.select("li.chapter")

    chapter_list = []
    for chap in chapters_elem:
        chap_title = chap.select_one("h4.chapter__title").getText().strip()
        chap_desc = chap.select_one("p.chapter__description").getText().strip()
        chap_details_elem = chap.select("h5.chapter__exercise-title")

        chap_detail_titles = []
        for cd in chap_details_elem:
            cd_title = cd.getText().strip()
            chap_detail_titles.append(cd_title)

        chapter_detail = {"title":chap_title, "desc":chap_desc, "details":chap_detail_titles}
        print(chapter_detail)
        chapter_list.append(chapter_detail)
    c["chapter_detail"] = chapter_list

wb = xl.Workbook()
sheet = wb.active
sheet.title = "Machine Learning Scientist with Python"

row = 2
col = 1
for c in courseList:
    sheet.cell(row=row, column=col).value = c["title"]
    sheet.cell(row=row, column=col+1).value = c["desc"]
    row += 1
    for cd in c["chapter_detail"]:
        sheet.cell(row=row, column=col+1).value = cd["title"]
        sheet.cell(row=row, column=col+2).value = cd["desc"]
        row += 1
        for cd_title in cd["details"]:
            sheet.cell(row=row, column=col+2).value = cd_title
            row += 1

wb.save("Machine Learning Scientist with Python.xlsx")

print(len(courseList))